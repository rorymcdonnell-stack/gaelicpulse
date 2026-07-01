"""
PULSE/26 — Canonical Data Access Layer
======================================
THE single source of truth is the master workbook (PULSE26_MASTER.xlsx),
'Master Fixtures' sheet. This module is the ONLY code that reads that
format. Everything else — book build, spreads, ledger, charts — imports
structured objects from here and never touches the spreadsheet directly.

One source of truth. No JSON. No drift.

Public API:
    load_master(path)        -> list[Game]
    games_by_competition(games)
    games_for_county(games, county)
    county_summaries(games)  -> dict[county] -> CountySummary
    all_scorers(games)       -> flat list of ScorerLine
    verify(games)            -> reconciliation report
"""
import re
from dataclasses import dataclass, field
from openpyxl import load_workbook

DEFAULT_PATH = '/home/claude/pulse/PULSE26_MASTER.xlsx'

# ── Data classes ──────────────────────────────────────────────────
@dataclass
class Scorer:
    player: str
    goals: int
    points: int
    tally: str
    note: str = ''
    @property
    def total(self): return self.goals*3 + self.points

@dataclass
class Game:
    idx: int
    date: str
    competition: str
    round: str
    venue: str
    team_a: str
    score_a: str
    team_b: str
    score_b: str
    two_a: int = None
    two_b: int = None
    aet: bool = False
    referee: str = ''
    attendance: str = ''
    scorers_a: list = field(default_factory=list)
    scorers_b: list = field(default_factory=list)
    both_sides: bool = False
    source: str = ''
    note: str = ''

    # Parsed score helpers
    def _gp(self, s):
        m = re.match(r'(\d+)-(\d+)', str(s or '').strip())
        return (int(m.group(1)), int(m.group(2))) if m else (None, None)
    @property
    def goals_a(self): return self._gp(self.score_a)[0]
    @property
    def points_a(self): return self._gp(self.score_a)[1]
    @property
    def goals_b(self): return self._gp(self.score_b)[0]
    @property
    def points_b(self): return self._gp(self.score_b)[1]
    @property
    def total_a(self):
        g,p = self._gp(self.score_a); return g*3+p if g is not None else None
    @property
    def total_b(self):
        g,p = self._gp(self.score_b); return g*3+p if g is not None else None
    @property
    def margin(self):
        if self.total_a is None or self.total_b is None: return None
        return self.total_a - self.total_b

# ── Scorer string parser ──────────────────────────────────────────
_SCORER_RE = re.compile(r'^(.+?)\s+(\d+)-(\d+)\s*(?:\((.+)\))?\s*$')

def parse_scorers(cell):
    """
    'David Clifford 0-04 (1tp, 0-01f); Tony Brosnan 0-04 (2tp); ...'
    -> [Scorer(...), ...]
    Robust to varied spacing and the '0-04' zero-padded style.
    """
    out = []
    if not cell or str(cell).strip().upper() in ('MISSING','NONE',''):
        return out
    for chunk in str(cell).split(';'):
        chunk = chunk.strip()
        if not chunk: continue
        m = _SCORER_RE.match(chunk)
        if not m:
            # Fallback: find any G-P pattern
            m2 = re.search(r'(\d+)-(\d+)', chunk)
            if m2:
                name = chunk[:m2.start()].strip()
                out.append(Scorer(name, int(m2.group(1)), int(m2.group(2)),
                                  f"{m2.group(1)}-{m2.group(2)}"))
            continue
        name = m.group(1).strip()
        g, p = int(m.group(2)), int(m.group(3))
        note = (m.group(4) or '').strip()
        out.append(Scorer(name, g, p, f"{g}-{p}", note))
    return out

# ── Loader ────────────────────────────────────────────────────────
def _cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v

def load_master(path=DEFAULT_PATH):
    wb = load_workbook(path, data_only=True)
    ws = wb['Master Fixtures']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    H = {name: i+1 for i, name in enumerate(headers)}

    def col(r, name):
        return _cell(ws, r, H[name]) if name in H else None

    games = []
    for r in range(2, ws.max_row+1):
        date = col(r, 'Date')
        if not date:
            continue
        aet_raw = col(r, 'AET')
        aet = bool(aet_raw) and str(aet_raw).strip().lower() in ('yes','true','1','y')
        both_raw = str(col(r, 'Both Sides?') or '').strip().lower()
        g = Game(
            idx=col(r, '#') or (r-1),
            date=str(date),
            competition=str(col(r, 'Competition') or ''),
            round=str(col(r, 'Round') or ''),
            venue=str(col(r, 'Venue') or ''),
            team_a=str(col(r, 'Team A') or ''),
            score_a=str(col(r, 'Score A') or ''),
            team_b=str(col(r, 'Team B') or ''),
            score_b=str(col(r, 'Score B') or ''),
            two_a=col(r, '2pt A'),
            two_b=col(r, '2pt B'),
            aet=aet,
            referee=str(col(r, 'Referee') or '').strip(),
            attendance=str(col(r, 'Attendance') or '').strip() if col(r,'Attendance') else '',
            scorers_a=parse_scorers(col(r, 'Team A Scorers')),
            scorers_b=parse_scorers(col(r, 'Team B Scorers')),
            both_sides=(both_raw == 'yes'),
            source=str(col(r, 'Source') or '').strip(),
            note=str(col(r, 'Notes') or '').strip(),
        )
        games.append(g)
    return games

# ── Derived views ─────────────────────────────────────────────────
COMP_GROUP_ORDER = {
    'Ulster':1,'Leinster':2,'Munster':3,'Connacht':4,
    'Sam Maguire':5,'Tailteann':6,'Other':9
}
def comp_group(competition):
    c = competition or ''
    for p in ['Ulster','Leinster','Munster','Connacht','Tailteann']:
        if p in c: return p
    if 'All-Ireland' in c or 'Senior Championship' in c or 'Sam Maguire' in c:
        return 'Sam Maguire'
    return 'Other'

def games_for_county(games, county):
    """All games involving a county, from that county's perspective."""
    out = []
    for g in games:
        if g.team_a == county:
            out.append(('A', g))
        elif g.team_b == county:
            out.append(('B', g))
    return out

def all_scorers(games):
    """Flat list of (game, county, opponent, Scorer)."""
    rows = []
    for g in games:
        for s in g.scorers_a:
            rows.append((g, g.team_a, g.team_b, s))
        for s in g.scorers_b:
            rows.append((g, g.team_b, g.team_a, s))
    return rows

# ── Verification ──────────────────────────────────────────────────
def verify(games):
    report = {'total_games': len(games), 'scorer_mismatches': [],
              'missing_referee': [], 'missing_attendance': [],
              'one_sided': []}
    for g in games:
        # Reconcile A
        if g.scorers_a:
            sg = sum(s.goals for s in g.scorers_a)
            sp = sum(s.points for s in g.scorers_a)
            if g.goals_a is not None and (sg != g.goals_a or sp != g.points_a):
                report['scorer_mismatches'].append(
                    (g.date, f"{g.team_a} v {g.team_b}", 'A',
                     f"score {g.score_a}", f"scorers {sg}-{sp}"))
        # Reconcile B
        if g.scorers_b:
            sg = sum(s.goals for s in g.scorers_b)
            sp = sum(s.points for s in g.scorers_b)
            if g.goals_b is not None and (sg != g.goals_b or sp != g.points_b):
                report['scorer_mismatches'].append(
                    (g.date, f"{g.team_a} v {g.team_b}", 'B',
                     f"score {g.score_b}", f"scorers {sg}-{sp}"))
        if not g.referee or g.referee.upper()=='MISSING':
            report['missing_referee'].append((g.date, f"{g.team_a} v {g.team_b}"))
        if not g.attendance or g.attendance.upper()=='MISSING':
            report['missing_attendance'].append((g.date, f"{g.team_a} v {g.team_b}"))
        if not (g.scorers_a and g.scorers_b):
            report['one_sided'].append((g.date, f"{g.team_a} v {g.team_b}"))
    return report

if __name__ == '__main__':
    games = load_master()
    print(f"Loaded {len(games)} games from master workbook")
    rep = verify(games)
    print(f"Scorer mismatches: {len(rep['scorer_mismatches'])}")
    for m in rep['scorer_mismatches'][:10]:
        print("  ", m)
    print(f"Missing referee: {len(rep['missing_referee'])}")
    print(f"Missing attendance: {len(rep['missing_attendance'])}")
    print(f"One-sided (scorers): {len(rep['one_sided'])}")
    total_scorers = sum(len(g.scorers_a)+len(g.scorers_b) for g in games)
    print(f"Total scorer records parsed: {total_scorers}")
